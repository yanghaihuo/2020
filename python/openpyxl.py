# -*- coding: utf-8 -*-
"""
Created on Thu May 16 13:10:25 2019

@author: Administrator
"""

from openpyxl import load_workbook
import os

os.chdir(r'C:\Users\Administrator\Desktop')
wb = load_workbook(r"优势电商企业出数模板包装指数.xlsx")


print(wb.sheetnames)


sheet = wb.get_sheet_by_name("交易指数")
#print(sheet["C"])
#print(sheet["4"])
#print(sheet['A2'])
#print(sheet['A2'].value)
#print(sheet.max_row)
#print(sheet.max_column)
#for i in sheet["C"]:
#  print(i.value, end=" ")



from openpyxl import Workbook

wb = Workbook()
sheet = wb.active #获取当前活跃的sheet,默认是第一个sheet
sheet = workbook.get_sheet_names()         #从名称获取sheet

sheet.title = "lianxi"




#存一行数据
sheet.append([11,87])


sheet['C3'] = 'Hello world!'
for i in range(10):
  sheet["A%d" % (i+1)].value = i + 1
  
  
sheet["E1"].value = "=SUM(A:A)"


wb.save('lianxi.xlsx')








#字体修改函数
from openpyxl.styles import Font          
fontobj = Font(name = 'Times New Roman' , size = 24,bold = True, italic = True)     #字体名称，字号，粗体，斜体
sheet['A3'].font =fontobj



sheet.row_dimensions[1].height =70   #修改行高
sheet.column_dimensions['B'].width = 20  #修改列宽


#图表
refobj = openpyxl.chart.Reference(sheet,(1,1),(10,1))              #创建Reference对象  (表格对象，左上角起始数据，右下角结束数据位置)
seriesobj = openpyxl.chart.Series(refobj,title = 'First Series')   #创建Series对象
chartobj = openpyxl.chart.BarChart()                               #图表类型 :  .lineChart  ScatterChar  PieChart 折线图、散点图、饼图
chartobj.append(seriesobj)                                         #添加数据构成图表


txtfiles = []    #找到当前目录所有的txt文件
for txt in os.listdir('.'):
    if txt.endswith('.txt'):
        txtfiles.append(txt)






df.to_csv("test_pd.csv", sep=',', header=False, index=False)





for sheetObj in Sheet.columns [1]：
       print（cellObj.value）

#创建sheet  
wb.create_sheet(index = 0，title ='First Sheet') 
#删除sheet
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))







import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
#合并单元格
sheet.merge_cells('A1:D3')
#取消合并单元格
sheet.unmerge_cells('A1：D3')

sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5')
sheet.unmerge_cells('C5 ：D5')
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')




#冻结窗格
freeze_panes 设置
行和列冻结
sheet.freeze_panes = 'A2'  第1行
sheet.freeze_panes = 'B1'  A栏
sheet.freeze_panes = 'C1'  A列和B列
sheet.freeze_panes = 'C2'  第1行和第A和B列
#取消冻结窗格
sheet.freeze_panes = 'A1' or sheet.freeze_panes = None



@@按行或列方式获取表中的数据
#sheet.rows，这是一个生成器，里面是每一行数据，每一行数据由一个元组类型包裹。
for row in worksheet.rows:
    for cell in row:
        print(cell.value,end=" ")
    print()
"""
各省市 工资性收入 家庭经营纯收入 财产性收入 转移性收入 食品 衣着 居住 家庭设备及服务 ……
北京市 5047.4 1957.1 678.8 592.2 1879.0 451.6 859.4 303.5 698.1 844.1 575.8 113.1 ……
天津市 3247.9 2707.4 126.4 146.3 1212.6 265.3 664.4 122.4 441.3 315.6 263.2 56.1 ……
……
"""
 
for col in worksheet.columns:
    for cell in col:
        print(cell.value,end=" ")
    print()
 
'''
各省市 北京市 天津市 河北省 山西省 内蒙古自治区 辽宁省 吉林省 黑龙江省 上海市 江苏省 浙江省 ……
工资性收入 5047.4 3247.9 1514.7 1374.3 590.7 1499.5 605.1 654.9 6686.0 3104.8 3575.1 ……
家庭经营纯收入 1957.1 2707.4 2039.6 1622.9 2406.2 2210.8 2556.7 2521.5 767.7 2271.4  ……
……
'''

@@获取特定行或特定列的数据
#输出特定的行
for cell in list(worksheet.rows)[3]:  #获取第四行的数据
    print(cell.value,end=" ")
print()
#河北省 1514.7 2039.6 107.7 139.8 915.5 167.9 531.7 115.8 285.7 265.4 166.3 47.0
 
#输出特定的列
for cell in list(worksheet.columns)[2]:  #获取第三列的数据
    print(cell.value,end=" ")
print()
#家庭经营纯收入 1957.1 2707.4 2039.6 1622.9 2406.2 2210.8 2556.7 2521.5 767.7 2271.4 3084.3……
 
#已经转换成list类型，自然是从0开始计数。


@@获取某一块的数据

for rows in list(worksheet.rows)[0:3]:
    for cell in rows[0:3]:
        print(cell.value,end=" ")
    print()
'''
各省市 工资性收入 家庭经营纯收入 
北京市 5047.4 1957.1 
天津市 3247.9 2707.4 
'''
 
for i in range(1, 4):
    for j in range(1, 4):
        print(worksheet.cell(row=i, column=j).value,end=" ")
    print()
'''
各省市 工资性收入 家庭经营纯收入 
北京市 5047.4 1957.1 
天津市 3247.9 2707.4 
'''

@@获取某一单元格的数据


















